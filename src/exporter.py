from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from models import ScoredInfluencer, Summary


OUTPUT_COLUMNS = [
    "score",
    "grade",
    "name",
    "username",
    "platform",
    "followers",
    "engagement_rate",
    "avg_views",
    "country",
    "email",
    "bio",
    "profile_url",
    "match_reasons",
    "warnings",
]


def scored_to_dataframe(scored: list[ScoredInfluencer]) -> pd.DataFrame:
    rows = [item.model_dump() for item in scored]
    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)
    return df[OUTPUT_COLUMNS].sort_values(by=["score"], ascending=False)


def summary_to_dataframe(summary: Summary) -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("total imported", summary.total_imported),
            ("total after dedupe", summary.total_after_dedupe),
            ("A count", summary.a_count),
            ("B count", summary.b_count),
            ("C count", summary.c_count),
            ("D count", summary.d_count),
            ("average score", summary.average_score),
            ("top countries", summary.top_countries),
            ("top platforms", summary.top_platforms),
        ],
        columns=["metric", "value"],
    )


def style_sheet(writer: pd.ExcelWriter, sheet_name: str) -> None:
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="F6E4CA")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)

    if sheet_name == "Filtered Influencers":
        grade_colors = {
            "A": "C6EFCE",
            "B": "D9EAF7",
            "C": "FFF2CC",
            "D": "F4CCCC",
        }
        for row in worksheet.iter_rows(min_row=2):
            grade_value = str(row[1].value or "")
            fill_color = grade_colors.get(grade_value[:1])
            if fill_color:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=fill_color)

    workbook.active = 0


def export_to_excel(scored: list[ScoredInfluencer], summary: Summary, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    filtered_df = scored_to_dataframe(scored)
    summary_df = summary_to_dataframe(summary)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        filtered_df.to_excel(writer, sheet_name="Filtered Influencers", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        style_sheet(writer, "Filtered Influencers")
        style_sheet(writer, "Summary")

